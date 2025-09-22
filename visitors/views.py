from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from .models import Visitor
from .forms import VisitorForm
import openpyxl
from openpyxl.utils import get_column_letter

@login_required
def visitor_create(request):
    if request.method == 'POST':
        form = VisitorForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('visitors:visitor_list')
    else:
        form = VisitorForm()
    return render(request, 'visitors/visitor_form.html', {'form': form})

@login_required
def visitor_list(request):
    q = request.GET.get('q', '').strip()
    visitor_qs = Visitor.objects.all().order_by('-created_at')
    if q:
        visitor_qs = visitor_qs.filter(
            Q(first_name__icontains=q) |
            Q(middle_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(mobile__icontains=q)
        )

    paginator = Paginator(visitor_qs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'q': q,
    }
    return render(request, 'visitors/visitor_list.html', context)

@login_required
def export_visitors_excel(request):
    q = request.GET.get('q', '').strip()
    visitor_qs = Visitor.objects.all().order_by('-created_at')
    if q:
        visitor_qs = visitor_qs.filter(
            Q(first_name__icontains=q) |
            Q(middle_name__icontains=q) |
            Q(last_name__icontains=q) |
            Q(mobile__icontains=q)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Visitors'

    headers = ['ID', 'First Name', 'Middle Name', 'Last Name', 'Mobile', 'Num People', 'Address', 'Created At']
    ws.append(headers)

    for v in visitor_qs:
        ws.append([
            v.id,
            v.first_name,
            v.middle_name,
            v.last_name,
            v.mobile,
            v.num_people,
            v.address,
            v.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        ])

    # Adjust column widths
    for i, col in enumerate(ws.columns, 1):
        max_length = 0
        column = get_column_letter(i)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=visitors.xlsx'
    wb.save(response)
    return response
